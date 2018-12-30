# -*- coding: GBK -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def combine():
    #wb_combine = Workbook()
    #ws_combine = wb_combine.create_sheet("combine")
    #ws_combine.title = "combine"

    wb = load_workbook('courses.xlsx')
    ws_students = wb['students']
    ws_time = wb['time']
    ws_combine = wb.create_sheet(title='combine')

    ws_combine.append(['创建时间','课程名称','学习人数','学习时间'])
    for i in range(2,ws_students.max_row + 1):
        for k in range(2,ws_time.max_row + 1):
            if ws_students.cell(row=i,column=2).value == ws_time.cell(row=k,column=2).value:
                ws_combine.append([ws_students.cell(row=i,column=1).value,
                    ws_students.cell(row=i,column=2).value,
                    ws_students.cell(row=i,column=3).value,
                    ws_time.cell(row=k,column=3).value])

    wb.save('courses.xlsx')


def split():
    wb = load_workbook('courses.xlsx')
    ws_combine = wb['combine']
    
    names = set()
    for i in range(2,ws_combine.max_row + 1):
        names.add(ws_combine.cell(row=i,column=1).value.year)

    header = ['创建时间','课程名称','学习人数','学习时间']
    
    for name in names:
        wb = Workbook()
        ws_name = wb.active
        ws_name.append(header)
        for i in range(2,ws_combine.max_row + 1):
            if ws_combine.cell(row=i,column=1).value.year == name:
                ws_name.append([ws_combine.cell(row=i,column=1).value,
                                ws_combine.cell(row=i,column=2).value,
                                ws_combine.cell(row=i,column=3).value,
                                ws_combine.cell(row=i,column=4).value])
        wb.save(str(name) + '.xlsx')    
        
    
if __name__ == '__main__':
    combine()
    split()
